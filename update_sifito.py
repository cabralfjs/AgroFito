#!/usr/bin/env python3
"""
update_sifito.py
────────────────
Exporta as 5 tabelas do SIFITO (Condições de Utilização) para JSON.
Cada tabela gera um ficheiro separado para carregamento lazy no site.

Dependências:
    pip install playwright openpyxl
    playwright install chromium
"""

import asyncio, json, os, sys, tempfile
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from playwright.async_api import async_playwright

TIMEOUT_MS  = 60_000
BASE_DIR    = Path(__file__).parent

# Mapeamento colunas Excel → chaves JSON (igual em todas as tabelas)
COL_MAP = {
    0:  "cultura",
    3:  "sit_particular",
    4:  "ambiente",
    5:  "inimigo",
    6:  "nome_cient",
    8:  "uso_menor",
    12: "produto",
    13: "autorizacao",
    14: "numero",
    15: "funcao",
    16: "substancia",
    17: "epoca",
    18: "tecnica",
    19: "num_max_intervalo",
    20: "concentracao",
    21: "vol_calda",
    22: "dose",
    23: "intervalo_seg",
    24: "restricoes",
    26: "validade",
    27: "limite_comerc",
    28: "limite_util",
}

# As 5 fontes — ordem, URL, estado e ficheiro de saída
SOURCES = [
    dict(
        url    = "https://sifito.dgav.pt/divulgacao/usos",
        estado = "Autorizada",
        key    = "AUT",
        output = "data_autorizadas.json",
    ),
    dict(
        url    = "https://sifito.dgav.pt/divulgacao/usoscanceladosvendapermitida",
        estado = "Cancelada — Venda Permitida",
        key    = "CVP",
        output = "data_canceladas_venda_permitida.json",
    ),
    dict(
        url    = "https://sifito.dgav.pt/divulgacao/usoscanceladosvendainterditautilizacaopermitida",
        estado = "Cancelada — Venda Interdita / Util. Permitida",
        key    = "CVIP",
        output = "data_canceladas_venda_interdita_util_permitida.json",
    ),
    dict(
        url    = "https://sifito.dgav.pt/divulgacao/usoscanceladosvendautilizacaointerdita",
        estado = "Cancelada — Venda e Util. Interditas",
        key    = "CVUI",
        output = "data_canceladas_venda_util_interditas.json",
    ),
]


async def download_xlsx(page, url: str) -> bytes:
    """Navega para url e clica no botão Exportar para Excel."""
    print(f"   🌐  {url}")
    # "load" em vez de "networkidle" — mais tolerante em páginas com muitos dados
    # Timeout de 3 minutos para tabelas grandes (~48k registos)
    await page.goto(url, timeout=180_000, wait_until="load")
    await page.locator(".k-grid").wait_for(timeout=180_000)

    btn = page.locator(
        "button",
        has=page.locator("span.k-button-text", has_text="Exportar para Excel")
    ).first
    await btn.wait_for(timeout=TIMEOUT_MS)

    async with page.expect_download(timeout=180_000) as dl_info:
        await btn.click()

    download = await dl_info.value
    print(f"   📥  {download.suggested_filename}")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        await download.save_as(tmp.name)
        data = Path(tmp.name).read_bytes()
        Path(tmp.name).unlink()

    return data


def xlsx_to_records(xlsx_bytes: bytes, estado: str) -> list[dict]:
    """Converte xlsx para lista de dicts, acrescentando o campo estado."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(xlsx_bytes)
        tmp_path = tmp.name

    wb   = load_workbook(tmp_path, read_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    Path(tmp_path).unlink()

    records = []
    for row in rows[2:]:          # linha 0=grupo, linha 1=cabeçalhos, linha 2+=dados
        obj = {"estado": estado}
        for idx, key in COL_MAP.items():
            val = row[idx] if idx < len(row) else None
            if val is None:
                obj[key] = ""
            elif isinstance(val, datetime):
                obj[key] = val.strftime("%Y-%m-%d")
            else:
                v = str(val).strip()
                obj[key] = "" if v == "-" else v
        records.append(obj)

    return records


def save_json(records: list[dict], output: str, date: str):
    path = BASE_DIR / output
    path.write_text(
        json.dumps({"date": date, "records": records},
                   ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8"
    )
    size_mb = path.stat().st_size / 1_048_576
    print(f"   💾  {output}  ({len(records):,} registos · {size_mb:.1f} MB)")


async def main():
    today = datetime.now().strftime("%d/%m/%Y")
    print("=" * 55)
    print("  SIFITO Updater — 5 tabelas")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 55)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(accept_downloads=True)
        page    = await context.new_page()

        for i, src in enumerate(SOURCES, 1):
            print(f"\n[{i}/5] {src['estado']}")
            try:
                xlsx_bytes = await download_xlsx(page, src["url"])
                records    = xlsx_to_records(xlsx_bytes, src["estado"])
                save_json(records, src["output"], today)
            except Exception as e:
                print(f"   ❌  Erro: {e}", file=sys.stderr)
                sys.exit(1)

        await browser.close()

    print("\n🎉  Todas as tabelas actualizadas com sucesso!")


if __name__ == "__main__":
    asyncio.run(main())
