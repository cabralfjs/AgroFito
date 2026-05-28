#!/usr/bin/env python3
"""
update_produtos.py — AgroFito
──────────────────────────────
Exporta as 4 tabelas de Produtos Fitofarmacêuticos do SIFITO.

Diferença face a update_sifito.py:
  • Linha 0 é o cabeçalho (sem linha de grupo)
  • Dados a partir da linha 1
  • Campo "Função" é dividido em funcao_curta + funcao_mecanismo
"""

import asyncio, json, sys, tempfile
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from playwright.async_api import async_playwright

TIMEOUT_MS = 60_000
BASE_DIR   = Path(__file__).parent

COL_MAP = {
    0:  "designacao",
    1:  "autorizacao",
    2:  "numero",
    3:  "titular",
    4:  "tipo_util",
    5:  "substancia",
    6:  "teor",
    7:  "percentagem",
    8:  "formulacao",
    # 9 tratado em separado (split por |)
    10: "classificacao",
    11: "frases",
    12: "baixo_risco",
    13: "cand_subs",
    14: "mpb",
    15: "data_autorizacao",
}

FUNC_TYPES = [
    'Fungicida','Inseticida','Herbicida','Acaricida','Moluscicida',
    'Nematodicida','Rodenticida','Fumigante','Regulador','Adjuvante',
    'Feromona','Bactericida','Algicida','Desinfestante','Repelente',
]

SOURCES = [
    dict(url="https://sifito.dgav.pt/divulgacao/produtos",
         estado="Autorizada", key="AUT", output="prod_autorizadas.json"),
    dict(url="https://sifito.dgav.pt/divulgacao/produtoscanceladosvendapermitida",
         estado="Cancelada — Venda Permitida", key="CVP",
         output="prod_canceladas_venda_permitida.json"),
    dict(url="https://sifito.dgav.pt/divulgacao/produtoscanceladosvendainterditautilizacaopermitida",
         estado="Cancelada — Venda Interdita / Util. Permitida", key="CVIP",
         output="prod_canceladas_venda_interdita_util_permitida.json"),
    dict(url="https://sifito.dgav.pt/divulgacao/produtoscanceladosvendautilizacaointerdita",
         estado="Cancelada — Venda e Util. Interditas", key="CVUI",
         output="prod_canceladas_venda_util_interditas.json"),
]


def extract_funcao_tipo(funcao_curta: str) -> str:
    up = funcao_curta.upper()
    for t in FUNC_TYPES:
        if t.upper() in up:
            return t
    words = funcao_curta.strip().split()
    return words[0].capitalize() if words else ""


async def download_xlsx(page, url: str) -> bytes:
    print(f"   🌐  {url}")
    await page.goto(url, timeout=180_000, wait_until="load")
    await page.locator(".k-grid").wait_for(timeout=180_000)
    btn = page.locator(
        "button",
        has=page.locator("span.k-button-text", has_text="Exportar para Excel")
    ).first
    await btn.wait_for(timeout=TIMEOUT_MS)
    async with page.expect_download(timeout=180_000) as dl_info:
        await btn.click()
    download = await dl_info.value
    print(f"   📥  {download.suggested_filename}")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        await download.save_as(tmp.name)
        data = Path(tmp.name).read_bytes()
        Path(tmp.name).unlink()
    return data


def xlsx_to_records(xlsx_bytes: bytes, estado: str) -> list[dict]:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(xlsx_bytes); tmp_path = tmp.name
    wb = load_workbook(tmp_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    Path(tmp_path).unlink()

    records = []
    for row in rows[1:]:          # linha 0 = cabeçalho, dados de 1 em diante
        obj = {"estado": estado}
        for idx, key in COL_MAP.items():
            val = row[idx] if idx < len(row) else None
            if val is None:
                obj[key] = ""
            elif isinstance(val, datetime):
                obj[key] = val.strftime("%Y-%m-%d")
            else:
                v = str(val).strip()
                obj[key] = "" if v in ("-", "- -", "--") else v
        # Dividir Função
        funcao_raw = str(row[9] or '').strip() if len(row) > 9 else ''
        parts = funcao_raw.split('|')
        obj['funcao_curta']    = parts[0].strip() if parts else ''
        obj['funcao_mecanismo']= parts[1].strip() if len(parts) > 1 else ''
        obj['funcao_tipo']     = extract_funcao_tipo(obj['funcao_curta'])
        records.append(obj)
    return records


def save_json(records, output, date):
    path = BASE_DIR / output
    path.write_text(
        json.dumps({"date": date, "records": records},
                   ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8"
    )
    sz = path.stat().st_size / 1_048_576
    print(f"   💾  {output}  ({len(records):,} registos · {sz:.1f} MB)")


async def main():
    today = datetime.now().strftime("%d/%m/%Y")
    print("=" * 55)
    print("  AgroFito — Produtos Fitofarmacêuticos (4 tabelas)")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 55)
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(accept_downloads=True)
        page    = await context.new_page()
        for i, src in enumerate(SOURCES, 1):
            print(f"\n[{i}/4] {src['estado']}")
            try:
                xlsx  = await download_xlsx(page, src["url"])
                recs  = xlsx_to_records(xlsx, src["estado"])
                save_json(recs, src["output"], today)
            except Exception as e:
                print(f"   ❌  Erro: {e}", file=sys.stderr)
                sys.exit(1)
        await browser.close()
    print("\n🎉  Produtos actualizados com sucesso!")

if __name__ == "__main__":
    asyncio.run(main())
